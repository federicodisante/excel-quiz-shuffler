import openpyxl
import random
import shutil
import os

def crea_quiz_con_numero(input_file, output_file):
    """
    Crea file di quiz randomizzati clonando il file di input e aggiungendo il numero del quiz.

    Args:
        input_file (str): Percorso del file Excel di input (.xlsx) contenente
                           il foglio "domande_risposte" e il foglio "template".
        output_file (str): Percorso del file Excel di output (.xlsx).
    """
    try:
        # Clona il file di input
        shutil.copyfile(input_file, output_file)

        # Apri il file clonato (il file di output)
        workbook_output = openpyxl.load_workbook(output_file)

        # Seleziona i fogli di lavoro
        domande_sheet = workbook_output["domande_risposte"]
        template_sheet = workbook_output["template"]

        if domande_sheet is None or template_sheet is None:
            print("Errore: Il file di input deve contenere i fogli 'domande_risposte' e 'template'.")
            workbook_output.close()
            os.remove(output_file)  # Rimuovi il file clonato in caso di errore
            return

        # Verifica che il foglio delle domande abbia il formato atteso
        if domande_sheet.max_row < 10 or domande_sheet.max_column < 5:
            print("Errore: Il foglio 'domande_risposte' non ha il formato previsto (almeno 10 righe x 5 colonne).")
            workbook_output.close()
            os.remove(output_file)  # Rimuovi il file clonato in caso di errore
            return

        domande_risposte = []
        for row in domande_sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=5):
            domanda = row[0].value
            risposte = [cell.value for cell in row[1:]]
            domande_risposte.append([domanda] + risposte)

        # Crea 20 fogli di lavoro basati sul template
        for i in range(20):
            numero_quiz = i + 1
            nome_foglio = f"Quiz_{numero_quiz}"
            nuovo_sheet = workbook_output.copy_worksheet(template_sheet)
            nuovo_sheet.title = nome_foglio

            # Inserisci il numero del quiz nella cella A46
            nuovo_sheet["A47"] = f"{numero_quiz}/20"

            # Mescola casualmente l'ordine delle domande
            domande_randomizzate = random.sample(domande_risposte, len(domande_risposte))

            # Inserisci le domande e le risposte riordinate nel nuovo foglio
            for j, qr in enumerate(domande_randomizzate):
                domanda = qr[0]
                risposte = qr[1:]
                random.shuffle(risposte)  # Mescola casualmente l'ordine delle risposte

                # Calcola l'indirizzo della domanda
                riga_domanda = 7 + (j * 4)
                indirizzo_domanda = f"B{riga_domanda}"
                nuovo_sheet[indirizzo_domanda] = domanda

                # Calcola l'indirizzo delle risposte
                riga_risposte = 9 + (j * 4)
                colonne_risposte = ["B", "D", "F", "H"]
                for k, risposta in enumerate(risposte):
                    indirizzo_risposta = f"{colonne_risposte[k]}{riga_risposte}"
                    nuovo_sheet[indirizzo_risposta] = risposta

        # Salva le modifiche nel file clonato (il file di output)
        workbook_output.save(output_file)
        print(f"Creato con successo il file: {output_file} con 20 fogli di quiz randomizzati (numero quiz aggiunto).")

    except FileNotFoundError:
        print(f"Errore: Il file '{input_file}' non è stato trovato.")
    except KeyError as e:
        print(f"Errore: Il foglio '{e}' non è stato trovato nel file di input.")
    except Exception as e:
        print(f"Si è verificato un errore: {e}")


file_input = "domande_risposte.xlsx"  # Sostituisci con il nome del tuo file di input
file_output = "quiz_randomizzati.xlsx"  # Sostituisci con il nome desiderato per il file di output

crea_quiz_con_numero(file_input, file_output)
